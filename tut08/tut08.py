
from datetime import datetime
start_time = datetime.now()

import os
import openpyxl
import pandas as pd


ind = open("india_inns2.txt","r+")                # Input files reading
pak = open("pak_inns1.txt","r+")
teams = open("teams.txt","r+")

team_ind_n_pak = teams.readlines()

pak_team = team_ind_n_pak[0]
pak_player = pak_team[23:-1:].split(",")

ind_team = team_ind_n_pak[2]
ind_player = ind_team[20:-1:].split(",")



ind_lines = ind.readlines()                       # Read the lines of input files
for i in ind_lines:
    if i == '\n':                                 # Removes extra line
        ind_lines.remove(i)
      
pak_lines = pak.readlines()                       # Same as for ind_lines
for i in pak_lines:
    if i == '\n':
        pak_lines.remove(i)

wb = openpyxl.Workbook()
sheet = wb.active

ind_wickets = 0                                   
pak_wickets = 0
ind_b = 0
pak_b = 0
total_bowler_ind = 0
total_bowler_pak = 0

ind_out={}
pak_out={}
ind_bowl={}
pak_bowl={}
ind_bat={}
pak_bat={}


for i in pak_lines:
    x = i.index(".")
    pak_over = i[0:x + 2]
    temp = i[x + 2::].split(",")
    cb = temp[0].split("to")

    if f"{cb[0].strip()}" not in ind_bowl.keys() :
        ind_bowl[f"{cb[0].strip()}"] = [1,0,0,0,0,0,0]

    elif "wide" in temp[1]:                                   # If ball was wide
        pass

    elif "bye" in temp[1]:                                    # adding runs
        if "FOUR" in temp[2]:
            pak_b += 4
        elif "1" in temp[2]:
            pak_b += 1
        elif "2" in temp[2]:
            pak_b += 2
        elif "3" in temp[2]:
            pak_b += 3
        elif "4" in temp[2]:
            pak_b += 4
        elif "5" in temp[2]:
            pak_b += 5

    else:
        ind_bowl[f"{cb[0].strip()}"][0] += 1
    
    if f"{cb[1].strip()}" not in pak_bat.keys() and temp[1] != "wide":
        pak_bat[f"{cb[1].strip()}"] = [0,1,0,0,0]

    elif "wide" in temp[1] :
        pass

    else:
        pak_bat[f"{cb[1].strip()}"][1] += 1
    
    if "out" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][3] += 1
        if "Bowled" in temp[1].split("!!")[0]:
            pak_out[f"{cb[1].strip()}"] = ("b" + cb[0])

        elif "Caught" in temp[1].split("!!")[0]:
            w = (temp[1].split("!!")[0]).split("by")
            pak_out[f"{cb[1].strip()}"] = ("c" + w[1] +" b " + cb[0])

        elif "Lbw" in temp[1].split("!!")[0]:
            pak_out[f"{cb[1].strip()}"] = ("lbw  b "+cb[0])

    if "no run" in temp[1] or "out" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 0
        pak_bat[f"{cb[1].strip()}"][0] += 0

    elif "1 run" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 1
        pak_bat[f"{cb[1].strip()}"][0] += 1

    elif "2 run" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 2
        pak_bat[f"{cb[1].strip()}"][0] += 2

    elif "3 run" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 3
        pak_bat[f"{cb[1].strip()}"][0] += 3

    elif "4 run" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 4
        pak_bat[f"{cb[1].strip()}"][0] += 4

    elif "FOUR" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 4
        pak_bat[f"{cb[1].strip()}"][0] += 4
        pak_bat[f"{cb[1].strip()}"][2] += 1

    elif "SIX" in temp[1]:
        ind_bowl[f"{cb[0].strip()}"][2] += 6
        pak_bat[f"{cb[1].strip()}"][0] += 6
        pak_bat[f"{cb[1].strip()}"][3] += 1

    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            ind_bowl[f"{cb[0].strip()}"][2] += int(temp[1][1])
            ind_bowl[f"{cb[0].strip()}"][5] += int(temp[1][1])

        else:
            ind_bowl[f"{cb[0].strip()}"][2] += 1
            ind_bowl[f"{cb[0].strip()}"][5] += 1


for i in pak_bat.values():
    i[-1]=round((i[0]/i[1])*100 , 2)


for i in ind_lines:
    x = i.index(".")
    ind_over=i[0:x+2]

    temp = i[x + 2::].split(",")

    cb = temp[0].split("to")
    if f"{cb[0].strip()}" not in pak_bowl.keys():
        pak_bowl[f"{cb[0].strip()}"]=[1,0,0,0,0,0,0]

    elif "wide" in temp[1]:
        pass

    elif "bye" in temp[1]:
        if "FOUR" in temp[2]:
            ind_b += 4
        elif "1" in temp[2]:
            ind_b += 1
        elif "2" in temp[2]:
            ind_b += 2
        elif "3" in temp[2]:
            ind_b += 3
        elif "4" in temp[2]:
            ind_b += 4
        elif "5" in temp[2]:
            ind_b += 5

    else:
        pak_bowl[f"{cb[0].strip()}"][0] += 1
    
    if f"{cb[1].strip()}" not in ind_bat.keys() and temp[1] != "wide":
        ind_bat[f"{cb[1].strip()}"]=[0,1,0,0,0]

    elif "wide" in temp[1] :
        pass

    else:
        ind_bat[f"{cb[1].strip()}"][1] += 1
    
    if "out" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][3] += 1
        
        if "Bowled" in temp[1].split("!!")[0]:
            ind_out[f"{cb[1].strip()}"] = ("b" + cb[0])

        elif "Caught" in temp[1].split("!!")[0]:
            w=(temp[1].split("!!")[0]).split("by")
            ind_out[f"{cb[1].strip()}"] = ("c" + w[1] +" b " + cb[0])

        elif "Lbw" in temp[1].split("!!")[0]:
            ind_out[f"{cb[1].strip()}"] = ("lbw  b "+ cb[0])

    if "no run" in temp[1] or "out" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 0
        ind_bat[f"{cb[1].strip()}"][0] += 0

    elif "1 run" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 1
        ind_bat[f"{cb[1].strip()}"][0] += 1

    elif "2 run" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 2
        ind_bat[f"{cb[1].strip()}"][0] += 2

    elif "3 run" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 3
        ind_bat[f"{cb[1].strip()}"][0] += 3

    elif "4 run" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 4
        ind_bat[f"{cb[1].strip()}"][0] += 4

    elif "FOUR" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 4
        ind_bat[f"{cb[1].strip()}"][0] += 4
        ind_bat[f"{cb[1].strip()}"][2] += 1

    elif "SIX" in temp[1]:
        pak_bowl[f"{cb[0].strip()}"][2] += 6
        ind_bat[f"{cb[1].strip()}"][0] += 6
        ind_bat[f"{cb[1].strip()}"][3] += 1

    elif "wide" in temp[1]:
        if "wides" in temp[1]:
            pak_bowl[f"{cb[0].strip()}"][2] += int(temp[1][1])
            pak_bowl[f"{cb[0].strip()}"][5] += int(temp[1][1])

        else:
            pak_bowl[f"{cb[0].strip()}"][2] += 1
            pak_bowl[f"{cb[0].strip()}"][5] += 1


for z in ind_bat.values():
    z[-1] = round((z[0]/z[1])*100 , 2)

for z in pak_bat.values():
    z[-1] = round((z[0]/z[1])*100 , 2)

for z in ind_bowl.values():
    if z[0]%6 == 0:
        z[0] = z[0]//6
    else:
        z[0] = (z[0]//6) + (z[0]%6)/10

for z in pak_bowl.values():
    if z[0]%6 == 0:
        z[0] = z[0]//6
    else:
        z[0] = (z[0]//6) + (z[0]%6)/10

for z in ind_bowl.values():
    x = str(z[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        z[-1] = round((z[2]/balls)*6,1)

    else:
        z[-1] = round((z[2]/z[0]) ,1) 


for z in pak_bowl.values():
    x = str(z[0])
    if "." in x:
        balls = int(x[0])*6 + int(x[2])
        z[-1] = round((z[2]/balls)*6,1)

    else:
        z[-1] = round((z[2]/z[0]) ,1)


pak_Batter = []
for key in pak_bat.keys():
    pak_Batter.append(key)

for i in range(len(pak_bat)):
    sheet.cell(5 + i, 1).value = pak_Batter[i]
    sheet.cell(5 + i, 5).value = pak_bat[pak_Batter[i]][0]
    sheet.cell(5 + i, 6).value = pak_bat[pak_Batter[i]][1]
    sheet.cell(5 + i, 7).value = pak_bat[pak_Batter[i]][2]
    sheet.cell(5 + i, 8).value = pak_bat[pak_Batter[i]][3]
    sheet.cell(5 + i, 9).value = pak_bat[pak_Batter[i]][4]

    if pak_Batter[i] not in pak_out:
        sheet.cell(5+i,3).value = "not out"
    else:
        sheet.cell(5+i,3).value=pak_out[pak_Batter[i]]

sheet.cell(3,1).value = "BATTERS"                                      # Just headings 
sheet["E3"] = "RUNS"
sheet["F3"] = "BALLS"
sheet["G3"] = " 4s "
sheet["H3"] = " 6s "
sheet["I3"] = "  SR  "


sheet["A18"] = "BOWLER"
sheet["C18"] = "OVER"
sheet["D18"] = "MAIDEN"
sheet["E18"] = "RUNS"
sheet["F18"] = "WICKET"
sheet["G18"] = "NO-BALL"
sheet["H18"] = "WIDE"
sheet["I18"] = "ECONOMY"

pak_bowlers = []
for key in pak_bowl.keys():
    pak_bowlers.append(key)

for i in range(len(pak_bowl)):                                          # input scores
    sheet.cell(42 + i, 1).value = pak_bowlers[i]
    sheet.cell(42 + i, 3).value = pak_bowl[pak_bowlers[i]][0]
    sheet.cell(42 + i, 4).value = pak_bowl[pak_bowlers[i]][1]
    sheet.cell(42 + i, 5).value = pak_bowl[pak_bowlers[i]][2]
    sheet.cell(42 + i, 6).value = pak_bowl[pak_bowlers[i]][3]
    sheet.cell(42 + i, 7).value = pak_bowl[pak_bowlers[i]][4]
    sheet.cell(42 + i, 8).value = pak_bowl[pak_bowlers[i]][5]
    sheet.cell(42 + i, 9).value = pak_bowl[pak_bowlers[i]][6]
    total_bowler_pak += pak_bowl[pak_bowlers[i]][2]
    ind_wickets += pak_bowl[pak_bowlers[i]][3]


sheet.cell(11 + len(pak_bat) + len(pak_bowl), 1).value = "# INDIA"
sheet.cell(11 + len(pak_bat) + len(pak_bowl), 2).value = " INNINGS"

ind_batters = []
for key in ind_bat.keys():
    ind_batters.append(key)

for i in range(len(ind_bat)):
    sheet.cell(31 + i, 1).value = ind_batters[i]
    sheet.cell(31 + i, 5).value = ind_bat[ind_batters[i]][0]
    sheet.cell(31 + i, 6).value = ind_bat[ind_batters[i]][1]
    sheet.cell(31 + i, 7).value = ind_bat[ind_batters[i]][2]
    sheet.cell(31 + i, 8).value = ind_bat[ind_batters[i]][3]
    sheet.cell(31 + i, 9).value = ind_bat[ind_batters[i]][4]

    if ind_batters[i] not in ind_out:
        sheet.cell(31 + i, 3).value = "not out"
    else:
        sheet.cell(31 + i, 3).value=ind_out[ind_batters[i]]

sheet["A29"] = "BATTERS"
sheet["E29"] = "RUNS"
sheet["F29"] = "BALLS"
sheet["G29"] = " 4s "
sheet["H29"] = " 6s "
sheet["I29"] = "  SR  "


sheet["A40"] = "BOWLER"
sheet["C40"] = "OVER"
sheet["D40"] = "MAIDEN"
sheet["E40"] = "RUNS"
sheet["F40"] = "WICKET"
sheet["G40"] = "NO-BALL"
sheet["H40"] = "WIDE"
sheet["I40"] = "ECONOMY"

ind_bowlers = []
for key in ind_bowl.keys():
    ind_bowlers.append(key)

for i in range(len(ind_bowl)):

    sheet.cell(20 + i, 1).value = ind_bowlers[i]
    sheet.cell(20 + i, 3).value = ind_bowl[ind_bowlers[i]][0]
    sheet.cell(20 + i, 4).value = ind_bowl[ind_bowlers[i]][1]
    sheet.cell(20 + i, 5).value = ind_bowl[ind_bowlers[i]][2]
    sheet.cell(20 + i, 6).value = ind_bowl[ind_bowlers[i]][3]
    sheet.cell(20 + i, 7).value = ind_bowl[ind_bowlers[i]][4]
    sheet.cell(20 + i, 8).value = ind_bowl[ind_bowlers[i]][5]
    sheet.cell(20 + i, 9).value = ind_bowl[ind_bowlers[i]][6]
    total_bowler_ind += ind_bowl[ind_bowlers[i]][2]
    pak_wickets += ind_bowl[ind_bowlers[i]][3]


ind_total_score = total_bowler_ind + pak_b                                   # Total Score
pak_total_score = total_bowler_pak + ind_b

sheet["E27"] = " " + str(ind_total_score) + " - " + str(ind_wickets)
sheet["F27"] = str(ind_over)
Eone = " " + str(pak_total_score) + " - " + str(pak_wickets)
Fone = str(pak_over)


wb.save("Scoreboard.xlsx")                                                   # Writing the pak_outut in csv format
df = pd.read_excel('Scoreboard.xlsx')
df = df.set_axis(['PAKISTAN', ' INNINGS'] + [" ", " ", Eone, Fone, " ", " ", " "], axis = 'columns')
df.to_csv('Scorecard.csv', index = False)


try:
    os.path.exists("Scoreboard.xlsx") 
    os.remove("Scoreboard.xlsx")

except:
    print("Unexpected Error!")



from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
