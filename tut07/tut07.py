import os
import math
import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
from datetime import datetime
start_time = datetime.now()


def oct_rank(mod, file_name):                                         # Calculates the Octants for all values
    oct_name_id_mapping = {
        "1":"Internal outward interaction",
        "-1":"External outward interaction", 
        "2":"External Ejection", 
        "-2":"Internal Ejection", 
        "3":"External inward interaction", 
        "-3":"Internal inward interaction", 
        "4":"Internal sweep", 
        "-4":"External sweep"
        }
    
    try:
        df = pd.read_excel(f'E:\\GitHub\\2001ME38_2022\\tut07\\input\\{file_name}')
        rows = df.shape[0]

    except Exception as e:
        print("File Was Not Found", e)


    U_mean = df['U'].mean()                                     # Calculates the values and average values of U', V' and W'
    V_mean = df['V'].mean()
    W_mean = df['W'].mean()

    df.insert(4, column = "U Avg", value = "")
    df.insert(5, column = "V Avg", value = "")
    df.insert(6, column = "W Avg", value = "")

    df.insert(7, column = "U' = U - U Avg", value = "")
    df.insert(8, column = "V' = V - V avg", value = "")
    df.insert(9, column = "W' = W - W avg", value = "")
    
    df["U' = U - U Avg"] = round(df['U'] - U_mean, 3)
    df["V' = V - V avg"] = round(df['V'] - V_mean, 3)
    df["W' = W - W avg"] = round(df['W'] - W_mean , 3)

    df.at[0, 'U Avg'] = round(U_mean, 3)
    df.at[0, 'V Avg'] = round(V_mean, 3)
    df.at[0, 'W Avg'] = round(W_mean, 3)

    df['U'] = round(df['U'], 3)
    df['V'] = round(df['V'], 3)
    df['W'] = round(df['W'], 3)
    df['T'] = round(df['T'], 3)

    df.insert(10, column = "Octant", value = "")
    df.insert(11, column = " ", value = "")
    df.insert(12, column = "", value = "")
    df.insert(13, column = "Octant ID", value = "")
    df.insert(14, column = "1", value = "")
    df.insert(15, column = "-1", value = "")
    df.insert(16, column = "2", value = "")
    df.insert(17, column = "-2", value = "")
    df.insert(18, column = "3", value = "")
    df.insert(19, column = "-3", value = "")
    df.insert(20, column = "4", value = "")
    df.insert(21, column = "-4", value = "")

    df.iloc[1, 12] = "Mod "+ str(mod)
    df.at[0, 'Octant ID'] = "Overall Count"
    list_1=[]

	
    for i in range(0, rows):                                                                 # Calculating Octants
        if df.at[i, "U' = U - U Avg"] >= 0 and  df.at[i, "V' = V - V avg"] >= 0: 
            if df.at[i, "W' = W - W avg"] < 0:
              df.at[i, 'Octant'] = -1
            else:
              df.at[i, 'Octant'] = 1

        elif df.at[i, "U' = U - U Avg"] < 0 and  df.at[i, "V' = V - V avg"] >= 0:
            if df.at[i, "W' = W - W avg"] < 0:
              df.at[i, 'Octant'] = -2
            else:
              df.at[i, 'Octant'] = 2

        elif df.at[i, "U' = U - U Avg"] < 0 and  df.at[i, "V' = V - V avg"] < 0:
            if df.at[i, "W' = W - W avg"] < 0:
              df.at[i, 'Octant'] = -3
            else:
              df.at[i, 'Octant'] = 3

        elif df.at[i, "U' = U - U Avg"] >= 0 and  df.at[i, "V' = V - V avg"] < 0:
            if df.at[i, "W' = W - W avg"] < 0:
              df.at[i, 'Octant'] = -4
            else:
              df.at[i, 'Octant'] = 4

        list_1.append(df.at[i, 'Octant'])

    df.at[0, "1"] = list_1.count(1)
    df.at[0, "-1"] = list_1.count(-1)
    df.at[0 ,"2"] = list_1.count(2)
    df.at[0 ,"-2"] = list_1.count(-2)
    df.at[0 ,"3"] = list_1.count(3)
    df.at[0 ,"-3"] = list_1.count(-3)
    df.at[0 ,"4"] = list_1.count(4)
    df.at[0 ,"-4"] = list_1.count(-4)
    
    list_start, indx = 0, 1
    list_end = len(list_1)

    st = int(mod)
    mod_rows_total = math.ceil(rows/st)
    rows_total = mod_rows_total


    for i in range(list_start, list_end, st):
        x = i
        list_2 = list_1[x : x + st]
        y = x + st - 1

        if y > rows:
            y = rows - 1
        df.at[indx ,'Octant ID'] = str(x)+"-"+str(y)
        df.at[indx, '1'] = list_2.count(1)
        df.at[indx, '-1'] = list_2.count(-1)
        df.at[indx, '2'] = list_2.count(2)
        df.at[indx, '-2'] = list_2.count(-2)
        df.at[indx, '3'] = list_2.count(3)
        df.at[indx, '-3'] = list_2.count(-3)
        df.at[indx, '4'] = list_2.count(4)
        df.at[indx, '-4'] = list_2.count(-4)
        indx = indx + 1

	
    
    try:                                                                                     # Adds octant rank

        col = 22
        
        for i in range(1,5):
            header = "Rank Octant " + str(i)
            df.insert(col, column = header, value = "")
            col = col + 1
            header = "Rank Octant " + str(-1*i)
            df.insert(col, column = header, value = "")
            col = col + 1

        df.insert(col, column = "Rank 1 Octant ID", value = "")
        col = col + 1
        df.insert(col, column = "Rank 1 Octant Name", value = "")
        col = col + 1
        
        dict = {}
        list_1 = []

        for i in range(1, 5):
            count_octant = df.at[0, str(i)]
            dict[count_octant] = str(i)
            list_1.append(count_octant)
            count_octant = df.at[0, str(-1*i)]
            dict[count_octant] = str(-1*i)
            list_1.append(count_octant)
        
        list_1.sort(reverse = True)
        rank = 1
        df.at[0, "Rank 1 Octant ID"] = dict[list_1[0]]
        df.at[0, "Rank 1 Octant Name"] = oct_name_id_mapping[dict[list_1[0]]]
        
        for i in list_1:
            octant_id = "Rank Octant "+dict[i]
            df.at[0, octant_id] = rank
            rank = rank + 1
        
        rank_1 = []

        for indx in range(1, mod_rows_total + 1): 
            dict = {}
            list_1 = []

            for i in range(1,5):
                count_octant = df.at[indx, str(i)]
                dict[count_octant] = str(i)
                list_1.append(count_octant)
                count_octant = df.at[indx, str(-1*i)]
                dict[count_octant] = str(-1*i)
                list_1.append(count_octant)

            list_1.sort(reverse = True)
            df.at[indx, "Rank 1 Octant ID"] = dict[list_1[0]]
            rank_1.append(dict[list_1[0]])
            df.at[indx, "Rank 1 Octant Name"] = oct_name_id_mapping[dict[list_1[0]]]
            
            rank = 1
            for i in list_1:
                octant_id = "Rank Octant " + dict[i]
                df.at[indx, octant_id] = rank
                rank = rank + 1
        
        indx = mod_rows_total + 5
        df.at[indx, 'Rank Octant 4'] = "Octant ID"
        df.at[indx, 'Rank Octant -4'] = "Octant Name"
        df.at[indx, 'Rank 1 Octant ID'] = "Count of Rank 1 Mod Values"
        indx = indx + 1


        for i in range(1, 5):
            octant_id = str(i)
            c = rank_1.count(octant_id)
            df.at[indx, 'Rank Octant 4'] = octant_id
            df.at[indx, 'Rank Octant -4'] = oct_name_id_mapping[octant_id]
            df.at[indx, 'Rank 1 Octant ID'] = c
            indx = indx + 1
            
            octant_id = str(-1*i)
            c = rank_1.count(octant_id)
            df.at[indx, 'Rank Octant 4'] = octant_id
            df.at[indx, 'Rank Octant -4'] = oct_name_id_mapping[octant_id]
            df.at[indx, 'Rank 1 Octant ID'] = c
            indx = indx + 1
            
    except Exception as e:
        print("Error! Not Expected", e)
            
    try:
        return df

    except:
        print("Error: fail to write to file!")

def octant_analysis(mod=5000):
	pass


def octant_transition(mod, df):                                                # Evaluates the transition values
    
    try:
        rows = df.shape[0]
        st = mod
        cols = df.shape[1]
        df.insert(cols, column = "                     ", value = "")

        cols = cols + 1

    except:
        print("Error: Input file cannot be read!")
    
    try:
        for i in range(2, 12):
            blank = ""

            for i in range(1, i + 1):
                blank += " "
            df.insert(cols, column = blank, value = "")
            cols = cols + 1

        dict_1 = {}
        blank_len = 4

        for i in range(1, 5):
            blank = ""
            for indx in range(0, blank_len):
                blank += " "

            dict_1[str(i)] = blank
            blank_len += 1
            blank = ""

            for indx in range(0, blank_len):
                blank += " "

            dict_1[str(-1*i)] = blank
            blank_len += 1

        dict_1['f'] = '  '
        dict_1['s'] = '   '
        
        indx = 0
        df.at[indx, dict_1['1']] = 'To'
        indx = indx + 1

        df.at[indx, dict_1['s']] = 'Count'
        for k in range(-4, 5):
            if k == 0:
                continue

            df.at[indx, dict_1[str(k)]] = k
        indx = indx + 1
        df.at[indx, dict_1['f']] = "From"

        data = []
        df_1 = pd.DataFrame(data, index = ['1','-1','2','-2','3','-3','4','-4'],
                        columns = ['1','-1','2','-2','3','-3','4','-4'])
        df_1 = df_1.fillna(0)

        for i in range(0, rows - 1):
            first = str(df.at[i, 'Octant'])
            second = str(df.at[i + 1, 'Octant'])
            df_1.at[first, second] += 1

        for i in range (1, 5):
            df.at[indx, dict_1['s']] = str(i)
            for j in range (-4, 5):
                if j == 0:
                    continue
                df.at[indx, dict_1[str(j)]] = df_1.at[str(i), str(j)]
            indx = indx + 1

            df.at[indx, dict_1['s']] = str(-1*i)
            for j in range (-4, 5):
                if j == 0:
                    continue
                df.at[indx, dict_1[str(j)]] = df_1.at[str(-1*i), str(j)]
            indx = indx + 1

    except Exception as e:
        print("Error! Not Expected", e)

    for i in range(0, rows, st):
        val = i + st
        if val >= rows:
            val = rows

        indx = indx + 2
        df.at[indx, dict_1['s']] = 'Mod Transition Count'
        indx = indx + 1

        df.at[indx, dict_1['s']] = str(i) + '-' + str(val - 1)
        df.at[indx, dict_1['1']] = 'To'
        indx = indx + 1

        df.at[indx, dict_1['s']] = 'Octant #'
        for k in range(-4, 5):
            if k == 0:
                continue
            df.at[indx, dict_1[str(k)]] =  k
        indx = indx + 1
        df.at[indx, dict_1['f']] = "From"

        data = []
        df_1 = pd.DataFrame(data, index = ['1','-1','2','-2','3','-3','4','-4'],
                        columns = ['1','-1','2','-2','3','-3','4','-4'])
        df_1 = df_1.fillna(0)

        if val == rows:
            val = val - 1

        for j in range(i, val):
            first = str(df.at[j,'Octant'])
            second = str(df.at[j+1, 'Octant'])
            df_1.at[first, second] += 1

        for k in range (1, 5):
            df.at[indx, dict_1['s']] = str(k)
            for l in range (-4, 5):
                if l == 0:
                    continue
                df.at[indx, dict_1[str(l)]] = df_1.at[str(k), str(l)]
            indx = indx + 1
            df.at[indx, dict_1['s']] = str(-1*k)

            for l in range (-4, 5):
                if l == 0:
                    continue
                df.at[indx, dict_1[str(l)]] = df_1.at[str(-1*k), str(l)]
            indx = indx + 1
    

    try:
        return df

    except Exception as e:    
        print("Error: Could not write to destination file!", e)

def octant_subsequence(mod, df, file_name):                                                      # Evaluates the longest subsequences for all values of octants

    try:
        rows = df.shape[0]
        cols = df.shape[1]
        df.insert(cols, column = "                   ", value = "")
        cols = cols + 1

    except Exception as e:
        print("Error: Cannot read input file!", e)
    
    try:
        data = []
        df_1 = pd.DataFrame(data, index=['1','-1','2','-2','3','-3','4','-4'],
                       columns=['Len', 'Count'])
        df_1 = df_1.fillna(0)
        
        df3 = pd.DataFrame(data, columns=['1','-1','2','-2','3','-3','4','-4'])
        
        prev = df.at[0, 'Octant']
        df_1.at[str(prev), 'Len'] = 1
        cur_l = 1

        a = df.at[0, 'T']
        b = df.at[0, 'T']

        for indx in range(1, rows):
            cur = df.at[indx, 'Octant']
            if (cur == prev):
                cur_l = cur_l + 1
            else:
                cur_l = 1
                a = df.at[indx, 'T']

            b = df.at[indx, 'T']
            df4 = df3.count(axis=0)

            if (cur_l == df_1.at[str(cur), 'Len']):
                df_1.at[str(cur), 'Count'] += 1                
                df3.at[df4[str(cur)], str(cur)] = a
                df3.at[df4[str(cur)]+1, str(cur)] = b

            elif(cur_l > df_1.at[str(cur), 'Len']):
                df_1.at[str(cur), 'Count'] = 1
                del df3[str(cur)]
                df3.insert(7, column = str(cur), value = "")
                df3.at[0, str(cur)] = a
                df3.at[1, str(cur)] = b

            df3.replace('', np.nan, inplace = True)
            df_1.at[str(cur), 'Len'] = max(cur_l, df_1.at[str(cur), 'Len'])
            prev = cur


        indx = 0

        for i in range(1, 5):
            df.at[indx, 'Octant ##'] = str(i)
            df.at[indx, 'Longest Subsequence Length'] = df_1.at[str(i), 'Len']
            df.at[indx, 'Count'] = df_1.at[str(i), 'Count']
            indx = indx + 1

            df.at[indx, 'Octant ##'] = str(-1*i)
            df.at[indx, 'Longest Subsequence Length'] = df_1.at[str(-1*i), 'Len']
            df.at[indx, 'Count'] = df_1.at[str(-1*i), 'Count']
            indx = indx + 1
            
        cols = df.shape[1]
        df.insert(cols, column = "                         ", value = "")


        indx = 0

        for i in range(1, 5):
            df.at[indx, 'Octant ###'] = str(i)
            df.at[indx, 'Longest Subsequence Length '] = df_1.at[str(i), 'Len']
            df.at[indx, 'Count '] = df_1.at[str(i), 'Count']
            indx = indx + 1

            df.at[indx, 'Octant ###'] = "Time"
            df.at[indx, 'Longest Subsequence Length '] = "From"
            df.at[indx, 'Count '] = "To"
            indx = indx + 1

            for index in range(0, len(df3[str(i)]), 2):
                if np.isnan(df3.at[index, str(i)]):
                    break

                df.at[indx, 'Longest Subsequence Length '] = df3.at[index, str(i)]
                df.at[indx, 'Count '] = df3.at[index+1, str(i)]
                indx = indx + 1
            
            df.at[indx, 'Octant ###'] = str(-1*i)
            df.at[indx, 'Longest Subsequence Length '] = df_1.at[str(-1*i), 'Len']
            df.at[indx, 'Count '] = df_1.at[str(-1*i), 'Count']
            indx = indx + 1

            df.at[indx, 'Octant ###'] = "Time"
            df.at[indx, 'Longest Subsequence Length '] = "From"
            df.at[indx, 'Count '] = "To"
            indx = indx + 1
            
            for index in range(0, len(df3[str(-1*i)]), 2):
                if np.isnan(df3.at[index, str(-1*i)]):
                    break

                df.at[indx, 'Longest Subsequence Length '] = df3.at[index, str(-1*i)]
                df.at[indx, 'Count '] = df3.at[index+1, str(-1*i)]
                indx = indx + 1

    except Exception as e:
        print("Error! Not Expected", e)
    
    try:
        df.to_excel(f'E:\\GitHub\\2001ME38_2022\\tut07\\output\\{file_name[0:len(file_name)-5]}_vel_octant_analysis_mod_{mod}.xlsx', index = False)
        return df

    except Exception as e:
        print("Error: Cannot write to output file!", e)


def octant_analysis(mod):                                                                            # Editing Output file and adding borders
 
	
    input_list = os.listdir('E:\\GitHub\\2001ME38_2022\\tut07\\input')                               # Listing input files

    for file_name in input_list:
        if(file_name[-4:] == 'xlsx'):
            
            print(file_name)
            df = oct_rank(mod, file_name)
            df = octant_transition(mod, df)
            df = octant_subsequence(mod, df, file_name)

             
        input_Path = 'E:\\GitHub\\2001ME38_2022\\tut07\\input\\' + file_name                                                                                                  # Input Path    
        output_Path = 'E:\\GitHub\\2001ME38_2022\\tut07\\output\\' + str(file_name[0:len(file_name)-5]) + '_vel_octant_analysis_mod_' + str(mod) + '.xlsx'                    # Output path

        w_s = xl.load_workbook(input_Path)
        sheet = w_s.active

		
        pattern_fill = PatternFill(patternType = "solid", fgColor = "FFFF33")                        # Edits files
        sheet['L1'].value = ""
        sheet['AG1'].value = ""
        sheet['AH1'].value = ""
        sheet['AR1'].value = ""

        tot_r = df.shape[0]
        tot_c = df.shape[1]
        rows_total = math.ceil(tot_r/mod)
        r = 0

        for row in sheet.iter_rows(min_row = 2, min_col = 1, max_row = tot_r + 1, max_col = tot_c):
            c = 0
            for cell in row: 
                cell.value = df.iat[r, c]
                c = c + 1
            r = r + 1

        for row in sheet.iter_rows(min_row = 1, min_col = 14, max_row = rows_total + 2, max_col = 32):
            for cell in row: 
                if(cell.value == 1):
                    cell.fill = pattern_fill

        sheet.column_dimensions['M'].width = 15
        sheet.column_dimensions['N'].width = 15
        sheet.column_dimensions['W'].width = 15
        sheet.column_dimensions['X'].width = 15
        sheet.column_dimensions['Y'].width = 15
        sheet.column_dimensions['Z'].width = 15
        sheet.column_dimensions['AA'].width = 15
        sheet.column_dimensions['AB'].width = 15
        sheet.column_dimensions['AC'].width = 15
        sheet.column_dimensions['AD'].width = 25
        sheet.column_dimensions['AE'].width = 24
        sheet.column_dimensions['AF'].width = 18
        sheet.column_dimensions['AT'].width = 24


        border_thin = Border(left=Side(border_style='thin',color='00000000'),
                        right=Side(border_style='thin',color='00000000'),
                        top=Side(border_style='thin',color='00000000'),
                        bottom=Side(border_style='thin',color='00000000')
                        )
        border_thick = Border(left=Side(border_style='thick',color='00000000'),
                    right=Side(border_style='thick',color='00000000'),
                    top=Side(border_style='thick',color='00000000'),
                    bottom=Side(border_style='thick',color='00000000')
                    )
                        
        col_n, r_loc, c_loc = 19, 1, 14

        for i in range(r_loc, r_loc + rows_total + 2):
            for j in range(c_loc, c_loc + col_n):
                sheet.cell(row = i, column = j).border = border_thick

        for i in range(rows_total + 7,rows_total + 16):
            for j in range(29, 32):
                sheet.cell(row = i, column = j).border = border_thick

        for row in sheet.iter_rows(min_row = 4, min_col = 36, max_row = 11, max_col = 43):
            for cell in row:
                cell.border = border_thick
        
        x = 3
        for n in range(rows_total+1):
            i = 0
            for row in sheet.iter_rows(min_row = x, min_col = 35, max_row = x + 8, max_col = 43):
                for cell in row:
                    if(cell.value != None):
                        cell.border = border_thick
            x = x + 13

        for row in sheet.iter_rows(min_row = 1, min_col = 45, max_row = 9, max_col = 47):
            for cell in row:
                cell.border = border_thick

        r_max = 1
        for row in range(2, tot_r):
            if str(sheet.cell(row = row, column = 50).value) == "nan":
                r_max = row
                break
        
        for row in sheet.iter_rows(min_row = 1, min_col = 49, max_row = r_max - 1, max_col = 51):
            for cell in row:
                cell.border = border_thick
        
        sheet['E1'] = 'U Avg'
        sheet['F1'] = 'V Avg'
        sheet['G1'] = 'W Avg'
        sheet['H1'] = "U' = U - U Avg"
        sheet['I1'] = "V' = V - V avg"
        sheet['J1'] = "W' = W - W avg"

        sheet['K1'] = 'Octant'
        sheet['N1'] = 'Overall Octant Count'
        sheet['O1'] = '+1'
        sheet['P1'] = '-1'
        sheet['Q1'] = '+2'
        sheet['R1'] = '-2'
        sheet['S1'] = '+3'
        sheet['T1'] = '-3'
        sheet['U1'] = '+4'
        sheet['V1'] = '-4'

        sheet['W1'] = 'Rank Octant 1'
        sheet['X1'] = 'Rank Octant -1'
        sheet['Y1'] = 'Rank Octant 2'
        sheet['Z1'] = 'Rank Octant -2'
        sheet['AA1'] = 'Rank Octant 3'
        sheet['AB1'] = 'Rank Octant -3'
        sheet['AC1'] = 'Rank Octant 4'
        sheet['AD1'] = 'Rank Octant -4'

        sheet['AE1'] = 'rank_1 Octant ID'
        sheet['AF1'] = 'rank_1 Octant Name'
        sheet['AI1'] = 'Overall Transition Count'
        sheet['AS1'] = 'Longest Subsquence Length'
        sheet['AW1'] = 'Longest Subsquence Length with Range'
        sheet['AX1'] = 'Longest Subsquence Length'
        sheet['AY1'] = 'Count'

		
        w_s.save(output_Path)                                             # Saves the output files


mod = 5000
octant_analysis(mod) 


##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 



from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")




#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
